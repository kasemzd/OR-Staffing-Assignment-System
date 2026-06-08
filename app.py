import streamlit as st
import pandas as pd
import urllib.parse
from io import BytesIO
import re

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

st.set_page_config(page_title="OR Rota", page_icon="🏥", layout="wide")

# =============================
# Styling
# =============================
st.markdown(
    """
    <style>
    .app-credit {
        color: #64748b;
        font-size: 0.70rem;
        margin-top: -12px;
        margin-bottom: 12px;
        font-weight: 500;
    }
    .section-title {
        text-align: left !important;
        margin-top: 4px;
        margin-bottom: 8px;
    }
    div.stButton > button,
    div.stDownloadButton > button,
    div[data-testid="stLinkButton"] > a {
        font-size: 17px !important;
        font-weight: 800 !important;
        padding: 0.65rem 1rem !important;
        border-radius: 10px !important;
        width: 100% !important;
        text-align: center !important;
    }
    .category-header {
        color: white;
        padding: 10px 14px;
        border-radius: 10px;
        font-weight: 900;
        font-size: 1.15rem;
        margin-top: 16px;
        margin-bottom: 10px;
    }
    .doctor-card {
        background-color:#0f172a;
        border:1px solid #1e293b;
        border-radius:10px;
        padding:12px 16px;
        margin-bottom:8px;
        display:flex;
        justify-content:space-between;
        align-items:center;
    }
    .doctor-card-room {
        font-weight:900;
        color:#38bdf8;
        font-size:17px;
    }
    .doctor-card-name {
        font-weight:900;
        color:#f8fafc;
        font-size:20px;
        text-align:right;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    # 🏥 Operation Room Rota
    ### Department of Anesthesia & Critical Care
    <div class="app-credit">Developed by Dr. Kasem</div>
    ---
    """,
    unsafe_allow_html=True,
)

# =============================
# Configuration
# =============================
MONTH_NAME = "June"
YEAR_NUMBER = 2026

OR_CATEGORIES = {
    "Main OR": ["OR1", "OR2", "OR4", "OR5", "OR8", "EMR OR", "OR10", "OR11", "OR12", "OR13", "OR16"],
    "PVT OR": ["OR14A", "OR14B", "14A (Sat)", "14B (Sat)", "14C (Sat)", "14D (Sat)"],
    "L&D": ["EMR CS"],
    "Basement": ["Base 1", "Base 2", "Base 3", "Base 4"],
    "Relievers": ["Rel. 1", "Rel. 2", "Rel. 3", "Rel (Sat)"],
    "PAC": ["PAC 1", "PAC 2", "PAC 3"],
    "On-Call 4 PM": ["OC 1", "OC 2"],
    "APS": ["APS"],
    "PACU": ["PACU"],
    "AOS": ["AOS 1", "AOS 2"],
    "1st Floor": ["A. Consultant"],
    "2nd Floor": ["2nd Floor A", "2nd Floor B"],
    "Backup": ["Backup 1", "Backup 2"],
}

CATEGORY_ORDER = list(OR_CATEGORIES.keys())

CATEGORY_ICONS = {
    "Main OR": "🏥",
    "PVT OR": "🔷",
    "L&D": "👶",
    "Basement": "⬇️",
    "Relievers": "🔁",
    "PAC": "🩺",
    "On-Call 4 PM": "📞",
    "APS": "💉",
    "PACU": "🛏️",
    "AOS": "👨‍⚕️",
    "1st Floor": "1️⃣",
    "2nd Floor": "2️⃣",
    "Backup": "🧯",
}

CATEGORY_COLORS = {
    "Main OR": "#2563eb",
    "PVT OR": "#7c3aed",
    "L&D": "#db2777",
    "Basement": "#475569",
    "Relievers": "#0891b2",
    "PAC": "#0f766e",
    "On-Call 4 PM": "#ea580c",
    "APS": "#16a34a",
    "PACU": "#ca8a04",
    "AOS": "#4f46e5",
    "1st Floor": "#0284c7",
    "2nd Floor": "#9333ea",
    "Backup": "#dc2626",
}

ROOM_ORDER = {room: order for category, rooms in OR_CATEGORIES.items() for order, room in enumerate(rooms)}

# =============================
# Helpers
# =============================
def get_category(room):
    for category, rooms in OR_CATEGORIES.items():
        if room in rooms:
            return category
    return "Other"


def clean_filename(text):
    text = str(text).strip()
    text = re.sub(r"[^A-Za-z0-9_-]+", "_", text)
    return text.strip("_") or "doctor"


def format_full_date(day_name, date_number):
    try:
        date_number = int(date_number)
    except Exception:
        pass
    return f"{day_name} {date_number} {MONTH_NAME} {YEAR_NUMBER}"


def prepare_export_df(df):
    export_df = df.copy()
    export_df["Date"] = export_df.apply(lambda r: format_full_date(r["Day"], r["Date"]), axis=1)
    export_df = export_df[["Date", "OR Category", "Room", "Doctor"]]
    return export_df


def prepare_rota(uploaded_file):
    wide_df = pd.read_excel(uploaded_file)

    required_cols = ["Date", "Day"]
    missing_cols = [col for col in required_cols if col not in wide_df.columns]
    if missing_cols:
        st.error(f"Missing required columns: {', '.join(missing_cols)}")
        st.stop()

    id_columns = ["Date", "Day"]
    room_columns = [col for col in wide_df.columns if col not in id_columns]

    long_df = wide_df.melt(
        id_vars=id_columns,
        value_vars=room_columns,
        var_name="Room",
        value_name="Doctor"
    )

    long_df = long_df.dropna(subset=["Doctor"])
    long_df = long_df[long_df["Doctor"].astype(str).str.strip() != ""]
    long_df = long_df[long_df["Doctor"].astype(str).str.lower() != "none"]

    invalid_doctors = ["14A", "14B", "14C", "14D", "REL", "PVT"]
    long_df = long_df[~long_df["Doctor"].astype(str).str.upper().isin(invalid_doctors)]

    saturday_map = {
        "Rel. 2": "14A (Sat)",
        "Rel. 3": "14B (Sat)",
        "PAC 1": "14C (Sat)",
        "PAC 2": "14D (Sat)",
        "PAC 3": "Rel (Sat)"
    }

    is_saturday = long_df["Day"].astype(str).str.lower().isin(["sa", "sat", "saturday"])
    mask = is_saturday & long_df["Room"].isin(saturday_map.keys())
    long_df.loc[mask, "Room"] = long_df.loc[mask, "Room"].map(saturday_map)
    long_df = long_df[~(is_saturday & (long_df["Room"] == "Rel. 1"))]

    long_df["Date"] = pd.to_numeric(long_df["Date"], errors="coerce")
    long_df = long_df.dropna(subset=["Date"])
    long_df["Date"] = long_df["Date"].astype(int)
    long_df["Day"] = long_df["Day"].astype(str)
    long_df["Room"] = long_df["Room"].astype(str)
    long_df["Doctor"] = long_df["Doctor"].astype(str)

    long_df["OR Category"] = long_df["Room"].apply(get_category)
    long_df["Category Order"] = long_df["OR Category"].apply(lambda x: CATEGORY_ORDER.index(x) if x in CATEGORY_ORDER else 999)
    long_df["Room Order"] = long_df["Room"].apply(lambda x: ROOM_ORDER.get(x, 999))

    return wide_df, long_df


def make_excel_download(df, sheet_name="OR Rota", title=None):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        startrow = 2 if title else 0
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31], startrow=startrow)
        worksheet = writer.sheets[sheet_name[:31]]

        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        if title:
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = title
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="left")

        header_row = startrow + 1
        header_fill = PatternFill("solid", fgColor="E2E8F0")
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

        for row in worksheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")

        for col_idx, column_cells in enumerate(worksheet.iter_cols(), start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 45)

    return output.getvalue()


def make_whatsapp_text(daily_df, daily_date, day_name):
    text = f"📅 {format_full_date(day_name, daily_date)}\n\n"
    for category in CATEGORY_ORDER:
        category_df = daily_df[daily_df["OR Category"] == category]
        if len(category_df) > 0:
            icon = CATEGORY_ICONS.get(category, "📌")
            text += f"{icon} {category}\n"
            for _, row in category_df.iterrows():
                text += f"{row['Room']} - {row['Doctor']}\n"
            text += "\n"
    return text.strip()


def add_pdf_footer(canvas, doc):
    canvas.saveState()
    width, height = A4
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.grey)
    canvas.drawCentredString(width / 2, 1.0 * cm, "Generated by OR Distribution Rota")
    canvas.restoreState()


def make_table_pdf(df, title, subtitle=""):
    if not REPORTLAB_AVAILABLE:
        return None

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.6 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TableTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        alignment=0,
        spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "TableSubtitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        alignment=0,
        spaceAfter=4,
    )
    credit_style = ParagraphStyle(
        "DevelopedCredit",
        parent=styles["Normal"],
        fontSize=7,
        textColor=colors.grey,
        alignment=0,
        spaceAfter=12,
    )

    export_df = df[["Date", "OR Category", "Room", "Doctor"]].copy()
    table_data = [list(export_df.columns)] + export_df.astype(str).values.tolist()

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[4.2 * cm, 3.7 * cm, 3.0 * cm, 6.0 * cm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    story = [Paragraph(title, title_style)]
    if subtitle:
        subtitle_clean = subtitle.replace("<br/>Developed by Dr. Kasem", "")
        story.append(Paragraph(subtitle_clean, subtitle_style))
        if "Developed by Dr. Kasem" in subtitle:
            story.append(Paragraph("Developed by Dr. Kasem", credit_style))
    story.append(table)

    doc.build(story, onFirstPage=add_pdf_footer, onLaterPages=add_pdf_footer)
    return buffer.getvalue()


def render_kpis(df, label="Doctor Duties"):
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("Doctors", df["Doctor"].nunique() if len(df) else 0)
    with k2:
        st.metric("Rooms", df["Room"].nunique() if len(df) else 0)
    with k3:
        st.metric(label, len(df))
    with k4:
        st.metric("Categories", df["OR Category"].nunique() if len(df) else 0)

# =============================
# App
# =============================
uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    wide_df, long_df = prepare_rota(uploaded_file)
    st.success(f"{len(wide_df)} days loaded")

    tab1, tab2, tab3 = st.tabs(["🔍 Search View", "📅 Daily Rota View", "👨‍⚕️ Doctor Duties"])

    # -----------------------------
    # Search View
    # -----------------------------
    with tab1:
        st.subheader("🔍 Search View")

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            selected_date = st.selectbox("Date", ["All"] + sorted(long_df["Date"].unique()), key="search_date")
        with col2:
            selected_category = st.selectbox("OR Category", ["All"] + CATEGORY_ORDER, key="search_category")

        category_filtered_df = long_df.copy()
        if selected_category != "All":
            category_filtered_df = category_filtered_df[category_filtered_df["OR Category"] == selected_category]

        with col3:
            selected_room = st.selectbox("Room", ["All"] + sorted(category_filtered_df["Room"].unique(), key=lambda x: ROOM_ORDER.get(x, 999)), key="search_room")
        with col4:
            selected_doctor = st.selectbox("Doctor", ["All"] + sorted(long_df["Doctor"].unique()), key="search_doctor")

        filtered_df = long_df.copy()
        if selected_date != "All":
            filtered_df = filtered_df[filtered_df["Date"] == selected_date]
        if selected_category != "All":
            filtered_df = filtered_df[filtered_df["OR Category"] == selected_category]
        if selected_room != "All":
            filtered_df = filtered_df[filtered_df["Room"] == selected_room]
        if selected_doctor != "All":
            filtered_df = filtered_df[filtered_df["Doctor"] == selected_doctor]

        filtered_df = filtered_df.sort_values(by=["Date", "Category Order", "Room Order"])
        result_df = prepare_export_df(filtered_df)

        render_kpis(filtered_df, label="Doctor Duties")

        has_search_filter = any([
            selected_date != "All",
            selected_category != "All",
            selected_room != "All",
            selected_doctor != "All",
        ])

        st.markdown("### Actions")
        if not has_search_filter:
            st.info("Please select at least one filter before exporting Excel or PDF.")

        a1, a2, a3, a4 = st.columns(4)
        with a1:
            st.download_button(
                "📊 Excel",
                data=make_excel_download(result_df, sheet_name="Search Results") if has_search_filter else b"",
                file_name="search_results_or_rota.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                disabled=not has_search_filter,
            )
        with a2:
            search_pdf = make_table_pdf(
                result_df,
                "Operation Room Rota",
                "Search Results<br/>Developed by Dr. Kasem",
            ) if has_search_filter else None
            if search_pdf:
                st.download_button(
                    "📑 PDF",
                    data=search_pdf,
                    file_name="search_results_or_rota.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    disabled=not has_search_filter,
                )
            else:
                st.download_button("📑 PDF", data=b"", file_name="search_results_or_rota.pdf", mime="application/pdf", use_container_width=True, disabled=True)
                if has_search_filter and not REPORTLAB_AVAILABLE:
                    st.warning("Install PDF package: pip install reportlab")
        with a3:
            st.empty()
        with a4:
            st.empty()

        st.markdown("---")
        st.subheader("Results")
        st.write(f"{len(result_df)} doctor duties found")
        st.dataframe(result_df, hide_index=True, width="stretch")

    # -----------------------------
    # Daily Rota View
    # -----------------------------
    with tab2:
        st.subheader("📅 Daily Rota View")

        daily_date = st.selectbox("Choose Date", sorted(long_df["Date"].unique()), key="daily_date")
        daily_df = long_df[long_df["Date"] == daily_date].copy()
        daily_df = daily_df.sort_values(by=["Category Order", "Room Order"])
        day_name = daily_df["Day"].iloc[0] if len(daily_df) > 0 else ""
        full_daily_date = format_full_date(day_name, daily_date)

        st.markdown(f'<h3 class="section-title">📅 {full_daily_date}</h3>', unsafe_allow_html=True)
        render_kpis(daily_df, label="Doctor Duties")

        whatsapp_text = make_whatsapp_text(daily_df, daily_date, day_name)
        wa_text = urllib.parse.quote(whatsapp_text)
        wa_link = f"https://wa.me/?text={wa_text}"
        daily_export_df = prepare_export_df(daily_df)

        st.markdown("### Actions")
        b1, b2, b3, b4 = st.columns(4)
        with b1:
            st.link_button("📱 WhatsApp", wa_link, use_container_width=True)
        with b2:
            st.download_button("📄 Text", data=whatsapp_text, file_name=f"rota_{daily_date}_{MONTH_NAME}_{YEAR_NUMBER}.txt", mime="text/plain", use_container_width=True)
        with b3:
            st.download_button(
                "📊 Excel",
                data=make_excel_download(daily_export_df, sheet_name="Daily Rota", title=f"Daily Rota - {full_daily_date}"),
                file_name=f"daily_rota_{daily_date}_{MONTH_NAME}_{YEAR_NUMBER}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with b4:
            daily_pdf = make_table_pdf(
                daily_export_df,
                "Operation Room Rota",
                f"Daily Rota - {full_daily_date}<br/>Developed by Dr. Kasem",
            )
            if daily_pdf:
                st.download_button("📑 PDF", data=daily_pdf, file_name=f"daily_rota_{daily_date}_{MONTH_NAME}_{YEAR_NUMBER}.pdf", mime="application/pdf", use_container_width=True)
            else:
                st.warning("Install PDF package: pip install reportlab")

        st.markdown("---")
        for category in CATEGORY_ORDER:
            category_df = daily_df[daily_df["OR Category"] == category]
            if len(category_df) > 0:
                icon = CATEGORY_ICONS.get(category, "📌")
                color = CATEGORY_COLORS.get(category, "#334155")
                st.markdown(f'<div class="category-header" style="background:{color};">{icon} {category}</div>', unsafe_allow_html=True)
                for _, row in category_df.iterrows():
                    st.markdown(
                        f"""
                        <div class="doctor-card">
                            <span class="doctor-card-room">{row['Room']}</span>
                            <span class="doctor-card-name">{row['Doctor']}</span>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

    # -----------------------------
    # Doctor Duties View
    # -----------------------------
    with tab3:
        st.subheader("👨‍⚕️ Doctor Duties")

        doctor_name = st.selectbox("Choose Doctor", sorted(long_df["Doctor"].unique()), key="doctor_schedule")
        doctor_df = long_df[long_df["Doctor"] == doctor_name].copy()
        doctor_df = doctor_df.sort_values(by=["Date", "Category Order", "Room Order"])
        doctor_export_df = prepare_export_df(doctor_df)
        safe_doctor = clean_filename(doctor_name)

        render_kpis(doctor_df, label="Doctor Duties")

        st.markdown("### Actions")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.download_button(
                "📊 Excel",
                data=make_excel_download(doctor_export_df, sheet_name="Doctor Duties", title=f"Doctor Duties - {doctor_name}"),
                file_name=f"doctor_duties_{safe_doctor}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with c2:
            doctor_pdf = make_table_pdf(
                doctor_export_df,
                "Operation Room Rota",
                f"Doctor Duties - {doctor_name}<br/>Developed by Dr. Kasem",
            )
            if doctor_pdf:
                st.download_button("📑 PDF", data=doctor_pdf, file_name=f"doctor_duties_{safe_doctor}.pdf", mime="application/pdf", use_container_width=True)
            else:
                st.warning("Install PDF package: pip install reportlab")
        with c3:
            st.empty()
        with c4:
            st.empty()

        st.markdown("---")
        st.markdown(f"### Doctor Duties - {doctor_name}")
        st.write(f"{len(doctor_df)} doctor duties found")
        st.dataframe(doctor_export_df, hide_index=True, width="stretch")

else:
    st.info("Please upload the monthly rota Excel file.")
